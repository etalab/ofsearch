# -*- coding: utf-8 -*-
import logging
import os
import sys

from urllib.request import urlopen, Request, urlretrieve

import click

from flask import Flask
from openpyxl import load_workbook

from .api import api
from .database import DB, DEFAULT_INDEX
from .utils import ObjectDict, is_tty


log = logging.getLogger(__name__)

CONTEXT_SETTINGS = {
    'help_option_names': ['-?', '-h', '--help']
}


def color(name, **kwargs):
    return lambda t: click.style(str(t), fg=name, **kwargs)


green = color('green', bold=True)
yellow = color('yellow', bold=True)
red = color('red', bold=True)
cyan = color('cyan')
magenta = color('magenta', bold=True)
white = color('white', bold=True)
bgred = color('white', bg='red')


OK = '✔'
KO = '✘'
WARNING = '⚠'
WAIT = '⏳'

PROGRESS_LABEL = ' '.join((cyan(WAIT), white('Loading organizations')))
DOWNLOAD_LABEL = ' '.join((cyan(WAIT), white('Downloading dataset')))
OPTIMIZE_LABEL = ' '.join((cyan(WAIT), white('Optimizing index using {memcpu}Mb on {cpus} cpu(s) (ie. {memory}Mb)')))


class ClickHandler(logging.Handler):
    '''Output using `click.echo`'''
    def emit(self, record):
        try:
            msg = self.format(record)
            level = record.levelname.lower()
            err = level in ('warning', 'error', 'exception', 'critical')
            click.echo(msg, err=err)
        except (KeyboardInterrupt, SystemExit):
            raise
        except:
            self.handleError(record)


class ClickFormatter(logging.Formatter):
    '''
    A log formatter using click ANSI colors and custom prefix when possible.
    '''
    LEVEL_COLORS = {
        'INFO': cyan,
        'WARNING': yellow,
        'ERROR': red,
        'CRITICAL': bgred,
        # 'DEBUG': bggrey
    }

    LEVEL_PREFIXES = {
        'INFO': cyan('ℹ'),
        'WARNING': yellow('⚠'),
        'ERROR': red('✘'),
        'CRITICAL': bgred('✘✘'),
    }

    def __init__(self, fmt=None, datefmt=None):
        fmt = fmt or '%(prefix)s %(message)s'
        super().__init__(fmt=fmt, datefmt=datefmt)

    def format_multiline(self, value, color):
        value = value.replace('\n', '\n{0} '.format(color('│')))
        # replace last by a folding char
        value = '╰'.join(value.rsplit('│', 1))
        return value

    def format(self, record):
        '''Customize the line prefix and indent multiline logs'''
        level_color = self.LEVEL_COLORS.get(record.levelname, white)
        std_prefix = '{0}:'.format(record.levelname)
        prefix = self.LEVEL_PREFIXES.get(record.levelname, std_prefix) if is_tty() else std_prefix
        record.__dict__['prefix'] = level_color(prefix)
        record.msg = self.format_multiline(record.msg, level_color)
        return super().format(record)

    def formatException(self, ei):
        '''Indent traceback info for better readability'''
        out = super().formatException(ei)
        out = red('│') + self.format_multiline(out, red)
        return out


@click.group(context_settings=CONTEXT_SETTINGS)
@click.option('-v', '--verbose', is_flag=True, help='Verbose output')
@click.option('-i', '--index', default=DEFAULT_INDEX, type=click.Path(writable=True),
              help='Index storage directory')
@click.pass_context
def cli(ctx, **kwargs):
    '''Elasticsearch loader for SIRENE dataset'''
    config = ctx.obj = ObjectDict(kwargs)

    log_level = logging.INFO if config.verbose else logging.WARNING

    handler = ClickHandler()
    handler.setLevel(log_level)
    handler.setFormatter(ClickFormatter())

    logger = logging.getLogger('ofsearch')
    logger.setLevel(log_level)
    logger.handlers = []
    logger.addHandler(handler)


@cli.command()
@click.argument('filename', type=click.Path())
@click.option('-m', '--memory', type=int, help="Limit memory usage (in Mb)", default=1024)
@click.pass_obj
def load(config, filename, memory):
    '''Load data from a official dataset file'''
    if filename.startswith('http://') or filename.startswith('https://'):
        filename = download_with_progress(filename)
    if not os.path.exists(filename):
        click.echo(' '.join([red(KO), white('Unable to find file {0}'.format(filename))]))
        sys.exit(1)
    wb = load_workbook(filename, read_only=True)
    sheet = wb.active  # Only the first sheet is relevant
    db = DB(config)
    total = sheet.max_row - 1
    with db.indexing(memory) as infos:
        with click.progressbar(sheet.rows, label=PROGRESS_LABEL, length=total) as rows:
            for i, row in enumerate(rows):
                if i == 0:
                    # This is the header row
                    fields = [cell.value for cell in row]
                else:
                    org = dict(zip(fields, [cell.value for cell in row]))
                    db.save_organization(org)
        click.echo(OPTIMIZE_LABEL.format(**infos))
    click.echo(green(OK) + white(' {0} items loaded with success'.format(i)))


def download_with_progress(url):
    req = Request(url, method='HEAD')
    req.add_header('Accept-Encoding', 'identity')
    response = urlopen(req)
    content_disposition = response.headers.get('Content-Disposition', '')
    if 'filename' in content_disposition:
        # Retrieve the filename and remove the last ".
        filename = content_disposition.split('filename="')[-1][:-1]
    else:
        filename = os.path.basename(url).strip()

    content_length = response.headers.get('Content-Length')
    if content_length:
        size = int(content_length)
    else:
        size = 1  # Fake for progress bar.

    with click.progressbar(length=size, label=DOWNLOAD_LABEL) as bar:
        def reporthook(blocknum, blocksize, totalsize):
            read = blocknum * blocksize
            if read <= 0:
                return
            if read >= totalsize:
                bar.update(size)
            else:
                bar.update(read)

        filename, _ = urlretrieve(url, filename, reporthook=reporthook)
        bar.update(size)
    click.echo(' '.join([green(OK), white('Downloaded file'), filename]))
    return filename


@cli.command()
@click.option('-d', '--debug', is_flag=True)
@click.option('--port', default=8888)
@click.pass_obj
def serve(config, debug, port):
    '''Launch a development server'''
    app = Flask(__name__)
    db = DB(config)
    app.config.SWAGGER_UI_DOC_EXPANSION = 'list'
    api.init_app(app)
    db.init_app(app)
    app.run(debug=debug, port=port)


@cli.command()
@click.pass_obj
def info(config):
    '''Display configuration and data statistics'''
    click.echo(cyan('OFSearch configuration'))
    for key, value in config.items():
        click.echo('{0}: {1}'.format(white(key), value))


@cli.command()
@click.pass_obj
def shell(config):
    '''Launch an interactive sheel (requires iPython)'''
    try:
        from IPython import embed
    except ImportError:
        log.error('This command requires ipython')
    db = DB(config)  # noqa: F841
    embed()


def main():
    '''
    Start the cli interface.
    This function is called from the entrypoint script installed by setuptools.
    '''
    cli(obj={}, auto_envvar_prefix='OFSEARCH')


# Allow running this file as standalone app without setuptools wrappers
if __name__ == '__main__':
    main()
